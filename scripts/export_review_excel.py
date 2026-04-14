"""
export_review_excel.py
Exports all test entries from rawData.json into a reviewer-friendly Excel workbook.
"""
import json
import re
import sys
sys.stdout.reconfigure(encoding='utf-8')

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ── helpers ────────────────────────────────────────────────────────────────────

def strip_html(text):
    """Remove HTML tags and decode common entities."""
    if not text:
        return ''
    # Replace <br> / <br/> with newline
    text = re.sub(r'<br\s*/?>', '\n', text, flags=re.IGNORECASE)
    # Replace links: keep inner text
    text = re.sub(r'<a[^>]*>(.*?)</a>', r'\1', text, flags=re.IGNORECASE | re.DOTALL)
    # Strip remaining tags
    text = re.sub(r'<[^>]+>', '', text)
    # Decode common HTML entities
    text = text.replace('&amp;', '&').replace('&lt;', '<').replace('&gt;', '>') \
               .replace('&nbsp;', ' ').replace('&#39;', "'").replace('&quot;', '"')
    return text.strip()

def get_plain_blocks(field):
    """Extract plain text from EditorJS paragraph/header blocks (joined by newline)."""
    if not field:
        return ''
    lines = []
    for block in field.get('blocks', []):
        btype = block.get('type', '')
        data = block.get('data', {})
        if btype in ('paragraph', 'header'):
            lines.append(strip_html(data.get('text', '')))
        elif btype == 'synonyms':
            syns = data.get('synonymList', [])
            if syns:
                lines.append(', '.join(syns))
        # other block types handled separately
    return '\n'.join(l for l in lines if l)

def get_synonyms(field):
    if not field:
        return ''
    for block in field.get('blocks', []):
        if block.get('type') == 'synonyms':
            return ', '.join(block['data'].get('synonymList', []))
    return ''

def get_alert_tags(field):
    """Return comma-separated alert tag labels."""
    if not field:
        return ''
    tags = []
    for block in field.get('blocks', []):
        if block.get('type') == 'alertTag':
            tags.append(block['data'].get('text', ''))
    return ', '.join(tags)

def get_lab_category(field):
    """Return list of (labName, categoryName) pairs."""
    if not field:
        return []
    pairs = []
    for block in field.get('blocks', []):
        if block.get('type') == 'lab_and_category':
            d = block.get('data', {})
            pairs.append((d.get('labName', ''), d.get('categoryName', '')))
    return pairs

def get_forms(field):
    """Return comma-separated form codes (and name if different)."""
    if not field:
        return ''
    forms = []
    for block in field.get('blocks', []):
        if block.get('type') == 'form':
            f = block['data'].get('form', {})
            code = f.get('form_code', '').strip()
            name = f.get('form_name', '').strip()
            only = f.get('formRequestOnly', False)
            if code:
                entry = code
                if name and name != code:
                    entry += ' (' + name + ')'
                if only:
                    entry += ' [form request only]'
                forms.append(entry)
    return '\n'.join(forms)

def get_container_text(field):
    """Return container description (text + container type)."""
    if not field:
        return ''
    lines = []
    for block in field.get('blocks', []):
        btype = block.get('type', '')
        data = block.get('data', {})
        if btype == 'paragraph':
            t = strip_html(data.get('text', ''))
            if t:
                lines.append(t)
        elif btype == 'container':
            # Use imageSrc filename as container type hint
            img = data.get('imageSrc', '')
            label = img.replace('/', '').replace('.webp', '').replace('_', ' ') if img else ''
            t = data.get('text', '').replace('##', '').replace('_', ' ').strip()
            desc = t or label
            if desc:
                lines.append('[Container: ' + desc + ']')
    return '\n'.join(lines)

# ── load data ──────────────────────────────────────────────────────────────────

with open('src/routes/rawData.json', 'r', encoding='utf-8') as f:
    raw = json.load(f)

tests = raw['testData']
alert_options = [''] + [a['text'] for a in raw['config'].get('alertTag', [])]
category_options = [''] + [c['text'] for c in raw['config'].get('category', [])]
lab_options = [''] + sorted(set(
    b['data'].get('labName', '')
    for t in tests
    for b in t.get('lab_and_category', {}).get('blocks', [])
    if b.get('type') == 'lab_and_category'
))

# ── Excel setup ────────────────────────────────────────────────────────────────

wb = openpyxl.Workbook()
wb.remove(wb.active)  # remove default sheet

# ── colours ───────────────────────────────────────────────────────────────────
HDR_BG   = PatternFill('solid', fgColor='1F3864')   # dark navy
HDR_FONT = Font(bold=True, color='FFFFFF', size=10)
SUB_BG   = PatternFill('solid', fgColor='D9E1F2')   # light blue subheader
SUB_FONT = Font(bold=True, color='1F3864', size=10)

HIDDEN_FILL   = PatternFill('solid', fgColor='E2E3E5')
SIMPLE_FILL   = PatternFill('solid', fgColor='FFF3CD')
REVIEWER_FILL = PatternFill('solid', fgColor='E2EFDA')

thin = Side(style='thin', color='CCCCCC')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

WRAP = Alignment(wrap_text=True, vertical='top')
CENTER = Alignment(horizontal='center', vertical='top', wrap_text=True)

# ── column definitions ─────────────────────────────────────────────────────────
# (header, width, description)
COLUMNS = [
    ('Test ID',              8,  'Numeric identifier of the test entry.'),
    ('CMS Name',             35, 'Name as it appears in the CMS ordering system (GCRS). Used as "Test Name" when Simple Layout is ON.'),
    ('Label Name',           30, 'Name printed on specimen labels. Same as CMS Name = Simple Layout candidate.'),
    ('Full Name (Legacy)',   35, 'Original long-form name. Scheduled for removal.'),
    ('Synonyms',             30, 'Alternative names / abbreviations (comma-separated). Used in search.'),
    ('Alert Tags',           35, 'Warning badges shown at the top of the test page (e.g. Special container, On-Ice transport).'),
    ('Lab',                  35, 'Laboratory responsible for performing this test.'),
    ('Category',             25, 'Test category used for sidebar filtering (e.g. Chemical Pathology, Haematology).'),
    ('Simple Layout',        14, 'YES = "Label Name" row is hidden from public view; "CMS Name" header becomes "Test Name". For tests where CMS Name = Label Name.'),
    ('Hidden',               10, 'YES = This entry is not shown to regular users (only visible in admin mode).'),
    ('Special Requirement',  50, 'Patient preparation, pre-test conditions, or special instructions before sample collection.'),
    ('Container / Sample',   35, 'Required specimen type and container (e.g. 4 mL clotted blood).'),
    ('Request Form(s)',       35, 'Paper/electronic request form required to accompany the specimen.'),
    ('Indications',          45, 'Clinical indications or when to order this test.'),
    ('Turnaround Time',      20, 'Expected time from receipt to result reporting.'),
    ('Lab Handling Notes',   50, 'Internal lab processing instructions (visible to lab staff and admin only).'),
    ('Reviewer Comments',    35, 'Free-text comments from reviewer. Leave blank if no issues.'),
    ('Review Status',        18, 'Select your assessment of this entry.'),
]

# ── Tests sheet ───────────────────────────────────────────────────────────────

ws = wb.create_sheet('Tests')

# Header row
for col_idx, (col_name, col_width, _) in enumerate(COLUMNS, 1):
    cell = ws.cell(row=1, column=col_idx, value=col_name)
    cell.fill = HDR_BG
    cell.font = HDR_FONT
    cell.alignment = CENTER
    cell.border = border
    ws.column_dimensions[get_column_letter(col_idx)].width = col_width

ws.row_dimensions[1].height = 30
ws.freeze_panes = 'A2'
ws.auto_filter.ref = ws.dimensions

# Data validations
dv_review = DataValidation(type='list', formula1='"OK,NEEDS EDIT,FLAG FOR DISCUSSION"', allow_blank=True, showDropDown=False)
ws.add_data_validation(dv_review)

# Write rows
for row_idx, test in enumerate(sorted(tests, key=lambda t: t['id']), 2):
    tid       = test['id']
    cms       = get_plain_blocks(test.get('GCRS_name'))
    label     = get_plain_blocks(test.get('label_name'))
    full      = get_plain_blocks(test.get('full_name'))
    synonyms  = get_synonyms(test.get('synonyms'))
    alerts    = get_alert_tags(test.get('alert'))
    lab_cats  = get_lab_category(test.get('lab_and_category'))
    labs      = '\n'.join(lc[0] for lc in lab_cats if lc[0])
    cats      = '\n'.join(set(lc[1] for lc in lab_cats if lc[1]))
    simple    = 'YES' if test.get('is_simple_layout') else 'NO'
    hidden    = 'YES' if test.get('is_hidden') else 'NO'
    req       = get_plain_blocks(test.get('requirement'))
    container = get_container_text(test.get('container'))
    forms     = get_forms(test.get('form'))
    indication= get_plain_blocks(test.get('indication'))
    tat       = get_plain_blocks(test.get('turn_around_time'))
    handling  = get_plain_blocks(test.get('test_handling'))

    row_values = [tid, cms, label, full, synonyms, alerts, labs, cats,
                  simple, hidden, req, container, forms, indication, tat,
                  handling, '', '']

    is_hidden_row  = test.get('is_hidden', False)
    is_simple_row  = test.get('is_simple_layout', False)

    for col_idx, value in enumerate(row_values, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.border = border
        cell.alignment = WRAP

        col_name = COLUMNS[col_idx - 1][0]

        # Row background for hidden entries
        if is_hidden_row:
            cell.fill = HIDDEN_FILL
        elif is_simple_row and col_name in ('Label Name', 'Simple Layout'):
            cell.fill = SIMPLE_FILL

        # Reviewer status column — centre + dropdown
        if col_name == 'Review Status':
            cell.alignment = CENTER
            dv_review.add(cell)
            cell.fill = REVIEWER_FILL

        if col_name == 'Reviewer Comments':
            cell.fill = REVIEWER_FILL

        # Simple Layout / Hidden — centre
        if col_name in ('Simple Layout', 'Hidden', 'Test ID'):
            cell.alignment = CENTER

    # Auto row height hint (openpyxl can't auto-size rows, set generous height)
    max_lines = max(
        (str(v).count('\n') + 1 if v else 1)
        for v in row_values
    )
    ws.row_dimensions[row_idx].height = min(max(15 * max_lines, 20), 200)

# ── Legend sheet ──────────────────────────────────────────────────────────────

wl = wb.create_sheet('Legend & Row Colours')
legend_data = [
    ('Row colour', 'Meaning'),
    ('Grey background', 'Entry is HIDDEN — not shown to regular users in the published site.'),
    ('Yellow in Label Name / Simple Layout', 'Simple Layout is ON — "Label Name" row is hidden in public view; "CMS Name" header shows as "Test Name".'),
    ('Green in last two columns', 'Reviewer columns — fill these in during review.'),
]
for r, (k, v) in enumerate(legend_data, 1):
    wl.cell(row=r, column=1, value=k).font = Font(bold=(r == 1))
    wl.cell(row=r, column=2, value=v)
wl.column_dimensions['A'].width = 40
wl.column_dimensions['B'].width = 70

wl.cell(row=6, column=1, value='Review Status values').font = Font(bold=True)
for i, val in enumerate(['OK', 'NEEDS EDIT', 'FLAG FOR DISCUSSION'], 7):
    wl.cell(row=i, column=1, value=val)

# ── Field Guide sheet ─────────────────────────────────────────────────────────

wg = wb.create_sheet('Field Guide')

guide_header = ['Column Name', 'Description']
for col_idx, h in enumerate(guide_header, 1):
    c = wg.cell(row=1, column=col_idx, value=h)
    c.fill = HDR_BG
    c.font = HDR_FONT
    c.alignment = CENTER

for row_idx, (col_name, _, description) in enumerate(COLUMNS, 2):
    wg.cell(row=row_idx, column=1, value=col_name).font = Font(bold=True)
    wg.cell(row=row_idx, column=2, value=description).alignment = Alignment(wrap_text=True, vertical='top')

wg.column_dimensions['A'].width = 22
wg.column_dimensions['B'].width = 80
wg.freeze_panes = 'A2'

# ── Alert Tags reference ──────────────────────────────────────────────────────

wg.cell(row=len(COLUMNS) + 3, column=1, value='Possible Alert Tags').font = Font(bold=True, color='CC0000')
for i, a in enumerate(raw['config'].get('alertTag', []), len(COLUMNS) + 4):
    wg.cell(row=i, column=1, value=a['text'])
    wg.cell(row=i, column=2, value='Badge colour: ' + a['background'])

# ── Labs reference ────────────────────────────────────────────────────────────

wg.cell(row=len(COLUMNS) + 3, column=4, value='All Laboratories').font = Font(bold=True, color='1F3864')
for i, lab in enumerate(lab_options[1:], len(COLUMNS) + 4):
    wg.cell(row=i, column=4, value=lab)

# ── Categories reference ──────────────────────────────────────────────────────

wg.cell(row=len(COLUMNS) + 3, column=6, value='All Categories').font = Font(bold=True, color='1F3864')
for i, cat in enumerate(raw['config'].get('category', []), len(COLUMNS) + 4):
    wg.cell(row=i, column=6, value=cat['text'] + '  (' + cat['code'] + ')')

# ── Save ──────────────────────────────────────────────────────────────────────

wb.sheet_order = ['Field Guide', 'Tests', 'Legend & Row Colours']

out_path = 'test_review.xlsx'
wb.save(out_path)
print('Saved to ' + out_path)
print('Total test entries: ' + str(len(tests)))
