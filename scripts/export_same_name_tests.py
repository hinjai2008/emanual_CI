import json
import re
import sys
sys.stdout.reconfigure(encoding='utf-8')

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

def get_text(field):
    """Extract plain text from an EditorJS field."""
    if not field:
        return ''
    blocks = field.get('blocks', [])
    if not blocks:
        return ''
    text = blocks[0].get('data', {}).get('text', '')
    # Strip HTML tags
    return re.sub(r'<[^>]+>', '', text).strip()

def get_lab_names(test):
    """Get comma-joined list of lab names from lab_and_category."""
    lac = test.get('lab_and_category')
    if not lac:
        return ''
    names = []
    for block in lac.get('blocks', []):
        ln = block.get('data', {}).get('labName', '')
        if ln:
            names.append(ln)
    return ', '.join(names)

with open('src/routes/rawData.json', 'r', encoding='utf-8') as f:
    data = json.load(f)

# Find tests where GCRS_name text == label_name text
matched = []
for test in data['testData']:
    gcrs = get_text(test.get('GCRS_name'))
    label = get_text(test.get('label_name'))
    if gcrs and label and gcrs == label:
        matched.append({
            'id': test['id'],
            'gcrs_name': gcrs,
            'label_name': label,
            'lab': get_lab_names(test),
            'current_simple_layout': test.get('is_simple_layout', False),
        })

matched.sort(key=lambda x: x['id'])
print(f"Found {len(matched)} tests with matching GCRS_name and label_name.")

# Build Excel
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Same Name Tests'

headers = ['Test ID', 'GCRS Name (= Label Name)', 'Lab', 'Current Simple Layout', 'Apply Simple Layout?']
header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF')

for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')

# Data validation for "Apply Simple Layout?" column (E)
dv = DataValidation(type='list', formula1='"YES,NO"', allow_blank=False)
ws.add_data_validation(dv)

yes_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
no_fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
already_fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')

for row_idx, entry in enumerate(matched, 2):
    ws.cell(row=row_idx, column=1, value=entry['id'])
    ws.cell(row=row_idx, column=2, value=entry['gcrs_name'])
    ws.cell(row=row_idx, column=3, value=entry['lab'])
    current_cell = ws.cell(row=row_idx, column=4, value='YES' if entry['current_simple_layout'] else 'NO')
    if entry['current_simple_layout']:
        current_cell.fill = already_fill

    # Default "Apply Simple Layout?" to YES if not already set, NO if already set
    default_apply = 'NO' if entry['current_simple_layout'] else 'YES'
    apply_cell = ws.cell(row=row_idx, column=5, value=default_apply)
    apply_cell.fill = yes_fill if default_apply == 'YES' else no_fill
    dv.add(apply_cell)

# Column widths
ws.column_dimensions['A'].width = 10
ws.column_dimensions['B'].width = 45
ws.column_dimensions['C'].width = 40
ws.column_dimensions['D'].width = 22
ws.column_dimensions['E'].width = 22

ws.freeze_panes = 'A2'
ws.auto_filter.ref = f'A1:E{len(matched) + 1}'

# Instructions sheet
ws2 = wb.create_sheet('Instructions')
ws2['A1'] = 'How to use this file'
ws2['A1'].font = Font(bold=True, size=13)
ws2['A2'] = '1. Review the "Same Name Tests" sheet.'
ws2['A3'] = '2. In column E ("Apply Simple Layout?"), set YES for tests you want to enable simple layout, NO to skip.'
ws2['A4'] = '3. Save the file and run: python scripts/apply_simple_layout_from_excel.py'
ws2['A5'] = ''
ws2['A6'] = 'Simple Layout effect: hides the Label Name row and relabels CMS Name as "Test Name" in public view.'
ws2.column_dimensions['A'].width = 80

out_path = 'same_name_tests.xlsx'
wb.save(out_path)
print(f"Saved to {out_path}")
