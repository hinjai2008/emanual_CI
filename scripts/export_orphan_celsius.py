import json, sys, re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding='utf-8')

with open('src/routes/rawData.json', encoding='utf-8') as f:
    data = json.load(f)

# °C not preceded by digit directly OR digit+space
bad_pattern = re.compile(r'(?<!\d)(?<!\d )\xb0C')

fields_to_check = ['requirement', 'test_handling']
FIELD_LABELS = {
    'requirement': 'Requirement',
    'test_handling': 'Test Handling',
}

rows = []
for entry in data['testData']:
    eid = entry['id']
    gcrs_blocks = (entry.get('GCRS_name') or {}).get('blocks', [])
    gcrs = next((b['data']['text'] for b in gcrs_blocks if (b.get('data') or {}).get('text')), '')

    for field in fields_to_check:
        for block in (entry.get(field) or {}).get('blocks', []):
            # Strip HTML tags for context display
            text = (block.get('data') or {}).get('text', '')
            text_plain = re.sub(r'<[^>]+>', '', text)
            for m in bad_pattern.finditer(text_plain):
                start = max(0, m.start() - 40)
                end = min(len(text_plain), m.end() + 40)
                snippet = ('...' if start > 0 else '') + text_plain[start:end].strip() + ('...' if end < len(text_plain) else '')
                rows.append({
                    'id': eid,
                    'gcrs_name': gcrs,
                    'field': FIELD_LABELS[field],
                    'snippet': snippet,
                })

# --- Build Excel ---
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Orphan °C Bugs'

# Styles
header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
header_fill = PatternFill(fill_type='solid', fgColor='1F4E79')
even_fill   = PatternFill(fill_type='solid', fgColor='EBF3FB')
odd_fill    = PatternFill(fill_type='solid', fgColor='FFFFFF')
warn_fill   = PatternFill(fill_type='solid', fgColor='FFF3CD')
center_align = Alignment(horizontal='center', vertical='center', wrap_text=False)
left_align   = Alignment(horizontal='left',   vertical='center', wrap_text=True)
thin = Side(style='thin', color='BDD7EE')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

headers = ['Test ID', 'GCRS Name', 'Field', 'Context (showing orphaned °C)']
col_widths = [10, 45, 18, 80]

# Header row
for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = border
    ws.column_dimensions[get_column_letter(col_idx)].width = width

ws.row_dimensions[1].height = 22

# Data rows
prev_id = None
fill_toggle = False
for row_idx, row in enumerate(rows, start=2):
    if row['id'] != prev_id:
        fill_toggle = not fill_toggle
        prev_id = row['id']

    row_fill = even_fill if fill_toggle else odd_fill

    values = [row['id'], row['gcrs_name'], row['field'], row['snippet']]
    for col_idx, value in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.fill = row_fill
        cell.border = border
        cell.font = Font(name='Calibri', size=10)
        if col_idx in (1,):
            cell.alignment = center_align
        else:
            cell.alignment = left_align

    # Highlight the snippet cell if it contains the orphan °C context
    snippet_cell = ws.cell(row=row_idx, column=4)
    snippet_cell.fill = warn_fill

ws.freeze_panes = 'A2'
ws.auto_filter.ref = ws.dimensions

out_path = 'orphan_celsius_bugs.xlsx'
wb.save(out_path)
sys.stdout.write('Saved: %s  (%d rows)\n' % (out_path, len(rows)))
